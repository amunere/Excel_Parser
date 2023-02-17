import openpyxl
from models import Excel_Data, db
 

class Parse_Excel:
    def get_excel(file: str):
        df = openpyxl.load_workbook(file)
        df_obj = df.active
        return df_obj
    
    def get_value(n, file):
        df_obj = Parse_Excel.get_excel(file)
        list = []
        for i in range(4, df_obj.max_row + 1):
            cell_obj = df_obj.cell(row = i, column = n)       
            list.append(cell_obj.value)
        return list
        
    def get_values(file):             
        val_list = []
        for i in range(2, 11):
            values = Parse_Excel.get_value(i, file)
            val_list.append(values)
        return val_list

    def load_db(file):
        load = Parse_Excel.get_values(file)
        val_company = load[0]
        val_fact_oliq_d1 = load[1]
        val_fact_oliq_d2 = load[2]
        val_fact_ooil_d1 = load[3]
        val_fact_ooil_d2 = load[4]
        val_forecast_oliq_d1 = load[5]
        val_forecast_oliq_d2 = load[6]
        val_forecast_ooil_d1 = load[7]
        val_forecast_ooil_d2 = load[8]
        d = []
        for i in range(0, 20):
            payload = Excel_Data(val_company[i], val_fact_oliq_d1[i], \
                val_fact_oliq_d2[i], val_fact_ooil_d1[i], val_fact_ooil_d2[i], \
                    val_forecast_oliq_d1[i], val_forecast_oliq_d2[i], \
                        val_forecast_ooil_d1[i], val_forecast_ooil_d2[i])
            payload.create()
        return payload